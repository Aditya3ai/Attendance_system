from flask import Flask, render_template, Response, request, redirect, url_for
import cv2
import os
import numpy as np
import openpyxl
import datetime

app = Flask(__name__)

# Path configurations
dataset_path = "datasets"  # Assuming 'datasets' is the folder name
excel_file = "attendance.xlsx"

# Initialize LBPH face recognizer
face_recognizer = cv2.face.LBPHFaceRecognizer_create()
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + "haarcascade_frontalface_default.xml")

# Load known faces and names
labels = {}
label_id = 0
training_data = []
training_labels = []

def load_dataset():
    global label_id
    for person_name in os.listdir(dataset_path):
        person_folder = os.path.join(dataset_path, person_name)
        if os.path.isdir(person_folder):
            for image_file in os.listdir(person_folder):
                image_path = os.path.join(person_folder, image_file)
                gray_image = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
                if gray_image is None:
                    continue

                faces = face_cascade.detectMultiScale(gray_image, scaleFactor=1.1, minNeighbors=5)
                for (x, y, w, h) in faces:
                    face_roi = gray_image[y:y + h, x:x + w]
                    training_data.append(face_roi)
                    training_labels.append(label_id)

            labels[label_id] = person_name
            label_id += 1

    # Train LBPH recognizer
    face_recognizer.train(training_data, np.array(training_labels))
    print("Dataset loaded and trained successfully!")

load_dataset()

# Initialize video capture
camera = cv2.VideoCapture(0)

def generate_frames():
    recognized_people = set()  # Track recognized people to avoid multiple entries in a short time
    while True:
        success, frame = camera.read()
        if not success:
            break
        else:
            gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
            faces = face_cascade.detectMultiScale(gray_frame, scaleFactor=1.1, minNeighbors=5)

            for (x, y, w, h) in faces:
                face_roi = gray_frame[y:y + h, x:x + w]
                label, confidence = face_recognizer.predict(face_roi)

                name = labels.get(label, "Unknown")
                if name != "Unknown" and name not in recognized_people:
                    add_attendance(name)  # Automatically add attendance
                    recognized_people.add(name)  # Mark this person as recognized

                # Draw bounding box and put text
                cv2.rectangle(frame, (x, y), (x + w, y + h), (0, 255, 0), 2)
                cv2.putText(frame, f"{name} ({round(confidence, 2)})", (x, y - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.9, (0, 255, 0), 2)

            ret, buffer = cv2.imencode('.jpg', frame)
            frame = buffer.tobytes()
            yield (b'--frame\r\n'
                   b'Content-Type: image/jpeg\r\n\r\n' + frame + b'\r\n')

@app.route('/')
def index():
    attendance_data = get_attendance_data()
    return render_template('index.html', attendance_data=attendance_data)

@app.route('/video_feed')
def video_feed():
    return Response(generate_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

def add_attendance(name):
    now = datetime.datetime.now()
    date = now.strftime("%Y-%m-%d")
    time = now.strftime("%H:%M:%S")
    try:
        if not os.path.exists(excel_file):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Name", "Date", "Time"])
        else:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active

        # Prevent duplicate attendance for the same person on the same day
        for row in ws.iter_rows(values_only=True):
            if row[0] == name and row[1] == date:
                print(f"Attendance for {name} already marked today.")
                return

        ws.append([name, date, time])
        wb.save(excel_file)
        print(f"Attendance for {name} added.")
    except Exception as e:
        print(f"Error saving attendance: {e}")

def get_attendance_data():
    data = []
    try:
        if os.path.exists(excel_file):
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                if row[0] != "Name":  # Skip header
                    data.append({'name': row[0], 'date': row[1], 'time': row[2]})
    except Exception as e:
        print(f"Error reading attendance: {e}")
    return data

def clear_attendance_data():
    try:
        if os.path.exists(excel_file):
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            ws.delete_rows(2, ws.max_row)  # Deletes all rows except the first (header row)
            wb.save(excel_file)
            print("Attendance data cleared successfully.")
        else:
            print("Attendance file does not exist.")
    except Exception as e:
        print(f"Error clearing attendance data: {e}")

@app.route('/clear_attendance', methods=['POST'])
def clear_attendance():
    clear_attendance_data()
    return redirect(url_for('index'))  # Redirect back to the home page

if __name__ == '__main__':
    app.run(debug=True)
